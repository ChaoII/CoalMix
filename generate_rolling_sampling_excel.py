# -*- coding: utf-8 -*-
"""生成"滚动采样{n}轮.xlsx"的可视化验证脚本（{n}=车数，由 --need-seq 决定）。

用法：
    python generate_rolling_sampling_excel.py
    python generate_rolling_sampling_excel.py --need-seq "5,6,5"
    python generate_rolling_sampling_excel.py --output "D:/输出目录"
    python generate_rolling_sampling_excel.py --output "D:/输出目录/自定义.xlsx"

参数：
    --need-seq  每车需求点数序列（逗号分隔），缺省用默认序列
    --output    输出路径。带 .xlsx 后缀视为文件路径；否则视为目录（自动创建，
                拼默认文件名）。缺省输出到 exe 同目录下的"滚动采样{n}轮.xlsx"。

依赖：
    openpyxl, src.auto_sampling（本项目）

输出：
    三个 sheet：编号映射 / 15车滚动 / 多批次随机性
"""
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 保证从任意目录运行都能 import 到 src
ROOT = Path(__file__).resolve().parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from src.auto_sampling import get_automatic_sampling_regions_rolling as roll
import src.auto_sampling as m

# 输出基础目录：优先"exe 同目录"（PyInstaller 打包时），否则脚本所在目录
if getattr(sys, "frozen", False):
    BASE_DIR = Path(sys.executable).resolve().parent
else:
    BASE_DIR = ROOT

# 默认输出文件名（--output 只给目录时使用），{n} = 车数（由 --need-seq 决定）
def default_out_name(num_rounds: int) -> str:
    return "滚动采样%d轮.xlsx" % num_rounds

# 默认的每车需求点数序列（可用命令行 --need-seq 覆盖）
DEFAULT_NEED_SEQ = [5, 6, 5, 6, 4, 7, 5, 6, 5, 5, 6, 4, 6, 5, 4]

# 初始化编号映射（列优先从右往左）
roll([], 1)
_NUMBERING = m._NUMBERING


def region_of(num: int) -> int:
    """编号 -> 大区（列0,1=大区0 最右 / 列2,3=大区1 / 列4,5=大区2 最左）。"""
    c = _NUMBERING[num][1]
    return 0 if c <= 1 else (1 if c <= 3 else 2)


def phase_of(num: int) -> str:
    """编号 -> 黑/白（棋盘相位）。"""
    r, c = _NUMBERING[num]
    return "黑" if (r + c) % 2 == 0 else "白"


def is_adj(a: int, b: int) -> bool:
    """两编号对应格子是否物理相邻（行差1同列 或 列差1同行）。"""
    r1, c1 = _NUMBERING[a]
    r2, c2 = _NUMBERING[b]
    return (abs(r1 - r2) == 1 and c1 == c2) or (abs(c1 - c2) == 1 and r1 == r2)


def sim_batch(need_seq):
    """按需求序列模拟 15 车滚动，记录每车明细。

    使用正确的 used 语义：换轮后前端传"新批次补足部分"。
    """
    rows = []
    used = []
    total = 18
    batch_order = None
    prev_last_region = None
    for i, need in enumerate(need_seq, 1):
        used_regions = [list(_NUMBERING[n]) for n in used]
        nums, cells = roll(used_regions, need)
        if batch_order is None:
            batch_order = list(m._BATCH_ORDER["default"])
        regs = [region_of(n) for n in nums]
        phases = [phase_of(n) for n in nums]
        unused = [n for n in range(1, total + 1) if n not in set(used)]
        wrapped = len(unused) < need
        link_ok = True
        if prev_last_region is not None:
            link_ok = (regs[0] - prev_last_region) % 3 == 1
        # 同车相邻判断：若发生换轮，nums = 旧批次收尾 + 新批次补足，
        # 这两段属于不同批次，各自内部判断相邻，跨段不判断。
        if wrapped:
            old_tail = nums[:len(unused)]
            new_fill = nums[len(unused):]
            adj_pairs = ([(old_tail[a], old_tail[b]) for a in range(len(old_tail))
                          for b in range(a + 1, len(old_tail)) if is_adj(old_tail[a], old_tail[b])]
                         + [(new_fill[a], new_fill[b]) for a in range(len(new_fill))
                            for b in range(a + 1, len(new_fill)) if is_adj(new_fill[a], new_fill[b])])
        else:
            adj_pairs = [(nums[a], nums[b]) for a in range(len(nums))
                         for b in range(a + 1, len(nums)) if is_adj(nums[a], nums[b])]
        has_dup = len(set(nums)) != len(nums)
        rows.append({
            "round": i, "need": need, "used": sorted(set(used)),
            "nums": nums, "cells": cells, "regs": regs, "phases": phases,
            "wrapped": wrapped, "link_ok": link_ok,
            "adj_pairs": adj_pairs, "has_dup": has_dup,
            "prev_last": prev_last_region,
        })
        prev_last_region = regs[-1]
        if wrapped:
            fill = need - len(unused)
            used = list(m._BATCH_ORDER["default"][:fill])
        else:
            used = sorted(set(used) | set(nums))
        if len(used) >= total:
            used = []
    return rows, batch_order


def build_workbook(need_seq=None) -> Workbook:
    wb = Workbook()
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", start_color="D9E1F2")
    warn_fill = PatternFill("solid", start_color="FCE4D6")
    bad_fill = PatternFill("solid", start_color="FFC7CE")
    ok_fill = PatternFill("solid", start_color="C6EFCE")
    r2_fill = PatternFill("solid", start_color="E2EFDA")
    r1_fill = PatternFill("solid", start_color="FFF2CC")
    r0_fill = PatternFill("solid", start_color="DDEBF7")
    hdr_font = Font(name="Arial", bold=True, size=11)
    body_font = Font(name="Arial", size=10)
    title_font = Font(name="Arial", bold=True, size=14)

    # ---- Sheet1: 编号-格子映射 ----
    ws = wb.active
    ws.title = "编号映射"
    ws.cell(row=1, column=1, value="编号规则：列优先从右往左（列0=最右，编号从最右列开始，每列内行0→行2）").font = title_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)

    ws.cell(row=3, column=1, value="网格视图（每格=编号，从左到右=列5→列0，列0在最右）").font = Font(name="Arial", bold=True, size=12)
    for c, h in enumerate(["", "列5", "列4", "列3", "列2", "列1", "列0"], 1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")
    for r in range(3):
        nums = []
        for c in range(5, -1, -1):
            num = [n for n, cell in _NUMBERING.items() if cell == (r, c)][0]
            nums.append(num)
        for c, v in enumerate(["行%d" % r] + nums, 1):
            cell = ws.cell(row=5 + r, column=c, value=v)
            cell.font = body_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
            if 2 <= c <= 7:
                col_idx = c - 2
                cell.fill = r0_fill if col_idx >= 4 else (r1_fill if col_idx >= 2 else r2_fill)
    ws.cell(row=9, column=1, value="绿=大区0(列0,1最右) | 黄=大区1(列2,3) | 蓝=大区2(列4,5最左)").font = body_font
    ws.merge_cells(start_row=9, start_column=1, end_row=9, end_column=7)

    ws.cell(row=11, column=1, value="编号明细（列优先从右往左）").font = Font(name="Arial", bold=True, size=12)
    for c, h in enumerate(["编号", "格子(row,col)", "所属大区", "列", "黑/白"], 1):
        cell = ws.cell(row=12, column=c, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.border = border
    for n in range(1, 19):
        r, c = _NUMBERING[n]
        for col, v in enumerate([n, "(%d,%d)" % (r, c), "大区%d" % region_of(n), "列%d" % c, phase_of(n)], 1):
            cell = ws.cell(row=12 + n, column=col, value=v)
            cell.font = body_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
    for col, w in zip("ABCDE", [8, 16, 12, 8, 8]):
        ws.column_dimensions[col].width = w

    # ---- Sheet2: 15车滚动 ----
    ws2 = wb.create_sheet("15车滚动")
    if need_seq is None:
        need_seq = DEFAULT_NEED_SEQ
    rows, batch_order = sim_batch(need_seq)
    headers = ["车次", "本车点数 need", "累计已用 used", "返回编号(批次序)",
               "对应大区序列", "黑/白序列", "同车相邻对", "编号重复?", "是否换轮", "衔接连续?"]
    widths = [6, 12, 26, 26, 26, 24, 20, 9, 8, 9]
    for c, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws2.cell(row=2, column=c, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws2.column_dimensions[get_column_letter(c)].width = w
    ws2.cell(row=1, column=1, value="15车滚动（大区全局0→1→2无缝连续；同车少相邻；无编号重复）").font = title_font
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)

    adj_count = 0
    dup_count = 0
    for i, r in enumerate(rows, 3):
        adj_desc = str(r["adj_pairs"]) if r["adj_pairs"] else "无"
        if r["adj_pairs"]:
            adj_count += 1
        if r["has_dup"]:
            dup_count += 1
        vals = [r["round"], r["need"], str(r["used"]), str(r["nums"]),
                str(r["regs"]), str(r["phases"]), adj_desc,
                "是" if r["has_dup"] else "否",
                "是" if r["wrapped"] else "否", "是" if r["link_ok"] else "否"]
        for c, v in enumerate(vals, 1):
            cell = ws2.cell(row=i, column=c, value=v)
            cell.font = body_font
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if c == 7 and r["adj_pairs"]:
                cell.fill = bad_fill
                cell.font = Font(name="Arial", size=10, bold=True, color="9C0006")
            if c == 7 and not r["adj_pairs"]:
                cell.fill = ok_fill
            if c == 8 and v == "是":
                cell.fill = bad_fill
            if c == 9 and v == "是":
                cell.fill = warn_fill
            if c == 10 and v == "否":
                cell.fill = bad_fill
    last = 2 + len(rows) + 2
    ws2.cell(row=last, column=1, value="本批次采样顺序: %s" % batch_order).font = body_font
    ws2.merge_cells(start_row=last, start_column=1, end_row=last, end_column=10)
    ws2.cell(row=last + 1, column=1,
             value="结果: 相邻车=%d, 重复车=%d（大区全程连续，仅need=7或换轮边界偶发相邻）"
                   % (adj_count, dup_count)).font = body_font
    ws2.merge_cells(start_row=last + 1, start_column=1, end_row=last + 1, end_column=10)

    # ---- Sheet3: 多批次随机性 ----
    ws3 = wb.create_sheet("多批次随机性")
    ws3.cell(row=1, column=1, value="多个独立批次（随机性保持：每批次18点顺序不同，同时窗口不相邻、大区连续）").font = title_font
    ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    h3 = ["批次", "采样顺序(18编号)", "大区序列", "前9色", "后9色"]
    for c, h in enumerate(h3, 1):
        cell = ws3.cell(row=2, column=c, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.border = border
    for b in range(1, 6):
        roll([], 1)
        order = list(m._BATCH_ORDER["default"])
        regs = [region_of(n) for n in order]
        vals = ["批次%d" % b, str(order), str(regs), phase_of(order[0]), phase_of(order[9])]
        for c, v in enumerate(vals, 1):
            cell = ws3.cell(row=2 + b, column=c, value=v)
            cell.font = body_font
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws3.cell(row=8, column=1, value="注：5个批次顺序互不相同（随机性）；每批次大区恒为[0,1,2]*6、前9同色后9另一色").font = body_font
    ws3.merge_cells(start_row=8, start_column=1, end_row=8, end_column=6)
    for col, w in zip("ABCDEF", [8, 44, 24, 8, 8, 8]):
        ws3.column_dimensions[col].width = w

    return wb, adj_count, dup_count


if __name__ == "__main__":
    import argparse
    import datetime

    parser = argparse.ArgumentParser(description="生成滚动采样 Excel 验证表")
    parser.add_argument(
        "--need-seq",
        type=str,
        default=None,
        help='每车需求点数序列，逗号分隔，如 "5,6,5,6,4,7"。缺省用默认序列 %s' % DEFAULT_NEED_SEQ,
    )
    parser.add_argument(
        "--output",
        type=str,
        default=None,
        help=('输出文件路径。可传目录或完整文件路径。缺省输出到 exe 同目录下的'
              '"滚动采样{n}轮.xlsx"（{n}=车数）'),
    )
    args = parser.parse_args()

    need_seq = DEFAULT_NEED_SEQ
    if args.need_seq:
        need_seq = [int(x.strip()) for x in args.need_seq.split(",") if x.strip()]
        if not need_seq:
            parser.error("--need-seq 需为逗号分隔的正整数，如 \"5,6,5\"")

    # 默认文件名随车数变化
    out_name = default_out_name(len(need_seq))

    # 解析输出路径：带 .xlsx 后缀视为文件路径，否则视为目录（拼默认文件名）
    if args.output:
        out_path = Path(args.output)
        is_dir = (out_path.is_dir() or str(args.output).endswith(("/", "\\"))
                  or out_path.suffix.lower() != ".xlsx")
        if is_dir:
            out_file = out_path / out_name
        else:
            out_file = out_path
    else:
        out_file = BASE_DIR / out_name

    # 确保输出目录存在
    out_file.parent.mkdir(parents=True, exist_ok=True)

    print("每车需求点数序列: %s" % need_seq)
    wb, adj_count, dup_count = build_workbook(need_seq)
    try:
        wb.save(out_file)
    except PermissionError:
        # 目标文件被占用（如正被 Excel 打开），输出带时间戳的新文件
        stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        out_file = out_file.with_name("%s_%s%s" % (out_file.stem, stamp, out_file.suffix))
        wb.save(out_file)
    print("saved %s" % out_file)
    print("相邻车=%d 重复车=%d" % (adj_count, dup_count))
